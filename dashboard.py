import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import plotly.graph_objects as go
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from datetime import date
from io import BytesIO
import base64
import os

# ─── PATHS ───────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ASSETS   = os.path.join(BASE_DIR, "assets")

# ─── PAGE CONFIG ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Custo de Produção – Aprosoja/MS",
    page_icon="🌱",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── HELPERS DE IMAGEM ───────────────────────────────────────────────────────
def _b64(path: str) -> str:
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode()

_logo_path  = os.path.join(ASSETS, "Aprosoja-4.png")
_fundo_path = os.path.join(ASSETS, "Fundo da pagina_clean.png")

apro_b64   = _b64(_logo_path)
fundo_b64  = _b64(_fundo_path) if os.path.exists(_fundo_path) else None
_WMARK_SRC = f"data:image/png;base64,{apro_b64}"


def _add_watermark(fig):
    fig.add_layout_image(dict(
        source=_WMARK_SRC,
        xref="paper", yref="paper",
        x=0.5, y=0.5,
        sizex=0.55, sizey=0.55,
        xanchor="center", yanchor="middle",
        opacity=0.08, layer="below",
    ))
    return fig


# ─── FONT FAMILY ─────────────────────────────────────────────────────────────
_font_face_css = ""
_font_path = os.path.join(ASSETS, "AprosojaMS.ttf")
if os.path.exists(_font_path):
    _fb64 = _b64(_font_path)
    _font_face_css = (
        "@font-face { font-family: 'AprosojaMS'; "
        f"src: url('data:font/truetype;base64,{_fb64}') format('truetype'); }}"
    )
_FONT = "'AprosojaMS', 'Segoe UI', Tahoma, Geneva, sans-serif"

# ─── FUNDO DA PÁGINA ─────────────────────────────────────────────────────────
if fundo_b64:
    st.markdown(f"""
    <style>
      [data-testid="stAppViewContainer"] {{
          background-image:
              linear-gradient(rgba(14,26,18,0.97), rgba(14,26,18,0.97)),
              url("data:image/png;base64,{fundo_b64}") !important;
          background-size: cover !important;
          background-attachment: fixed !important;
      }}
      [data-testid="stMain"] {{ background: transparent !important; }}
    </style>""", unsafe_allow_html=True)

# ─── CSS TEMA APROSOJA ────────────────────────────────────────────────────────
st.markdown(f"""
<style>
  {_font_face_css}

  html, body, [data-testid="stAppViewContainer"], [data-testid="stMain"],
  [data-testid="stSidebar"], [data-testid="stMarkdownContainer"],
  [data-testid="metric-container"], p, div.stApp, label,
  h1, h2, h3, h4, h5, h6 {{
      font-family: {_FONT} !important;
  }}

  [data-testid="stAppViewContainer"] {{ background-color: #0e1a12; }}
  [data-testid="stMain"]             {{ background-color: #0e1a12; }}
  .block-container                   {{ padding-top: 0 !important; }}

  /* Sidebar */
  [data-testid="stSidebar"] > div:first-child {{
      background: linear-gradient(180deg,#081410 0%,#0c1e16 60%,#081410 100%) !important;
      border-right: 1px solid #1c3d28 !important;
  }}
  [data-testid="stSidebar"] label,
  [data-testid="stSidebar"] p,
  [data-testid="stSidebar"] span,
  [data-testid="stSidebar"] .stMarkdown {{ color: #e8f5d0 !important; }}
  [data-testid="stSidebar"] hr {{ border-color: #1c3d28 !important; }}

  /* Multiselect sidebar */
  [data-testid="stSidebar"] [data-baseweb="select"] > div:first-child {{
      background-color: #1a3025 !important;
      border-color: #2d5a3d !important;
      color: #e8f0eb !important;
  }}
  [data-testid="stSidebar"] [data-baseweb="select"] span {{ color: #e8f0eb !important; }}
  [data-testid="stSidebar"] [data-baseweb="tag"] {{
      background-color: #1a3025 !important;
      border: 1px solid #2d5a3d !important;
  }}
  [data-testid="stSidebar"] [data-baseweb="tag"] span {{ color: #c8e6a0 !important; }}

  /* Botões cultura sidebar */
  [data-testid="stSidebar"] [data-testid="stHorizontalBlock"] [data-testid="baseButton-secondary"] {{
      background: #1a3025 !important; color: #c8e6a0 !important;
      border: 1px solid #2d5a3d !important; border-radius: 14px !important;
      font-weight: 600 !important; width: 100% !important;
  }}
  [data-testid="stSidebar"] [data-testid="stHorizontalBlock"] [data-testid="baseButton-primary"] {{
      background: #1a3025 !important; color: #ffffff !important;
      border: 2px solid #4caf50 !important;
      box-shadow: 0 0 10px rgba(76,175,80,0.35) !important;
      border-radius: 14px !important; font-weight: 700 !important; width: 100% !important;
  }}

  /* KPI containers */
  [data-testid="metric-container"],
  [data-testid="stMetric"] {{
      background: linear-gradient(145deg, #122518 0%, #1a3328 100%) !important;
      border: 1px solid #243f2f !important;
      border-radius: 12px !important;
      padding: 14px 18px !important;
      box-shadow: 0 6px 24px rgba(0,0,0,0.45) !important;
  }}
  [data-testid="stMetricLabel"],
  [data-testid="stMetricLabel"] * {{
      color: #6a9978 !important;
      font-size: 0.68rem !important;
      font-weight: 700 !important;
      text-transform: uppercase !important;
      letter-spacing: 0.9px !important;
  }}
  [data-testid="stMetricValue"],
  [data-testid="stMetricValue"] > div {{
      color: #e0ebe4 !important;
      font-size: 1.5rem !important;
      font-weight: 800 !important;
  }}
  [data-testid="stMetricDelta"] {{
      color: #5ed67f !important;
      font-size: 0.75rem !important;
  }}
  [data-testid="stMetricDelta"] svg {{ display: none !important; }}

  /* Texto geral */
  [data-testid="stMarkdownContainer"] p,
  [data-testid="stMarkdownContainer"] span,
  .stMarkdown p {{ color: #e0ebe4 !important; }}
  [data-testid="stCaptionContainer"] p {{ color: #7aa88a !important; }}

  h2, h3 {{
      color: #e8f0eb !important;
      border-left: 5px solid #c8a415;
      padding-left: 10px;
  }}

  /* Gráficos e tabelas */
  [data-testid="stPlotlyChart"] {{
      background: rgba(15,31,21,0.9) !important;
      border-radius: 16px !important;
      overflow: hidden !important;
      box-shadow: 0 4px 16px rgba(0,0,0,0.35) !important;
      border: 1px solid #1c3d28 !important;
  }}
  [data-testid="stDataFrame"] {{
      background: rgba(18,37,24,0.9) !important;
      border-radius: 16px !important;
      overflow: hidden !important;
      box-shadow: 0 4px 16px rgba(0,0,0,0.35) !important;
  }}
  [data-testid="stDataFrame"] th {{
      background-color: #2D5416 !important;
      color: #e8f5d0 !important;
      font-weight: 600 !important;
  }}

  hr {{ border-color: #1c3d28 !important; border-width: 1.5px !important; }}

  /* Card título de seção */
  .card-titulo {{
      background: linear-gradient(145deg, #122518 0%, #1a3328 100%);
      border: 1px solid #243f2f;
      border-radius: 12px;
      padding: 9px 16px;
      box-shadow: 0 2px 6px rgba(0,0,0,0.30);
      margin-bottom: 6px;
  }}
  .card-titulo h3 {{
      color: #ffffff !important;
      border-left: 4px solid #C4A040 !important;
      padding-left: 10px !important;
      margin: 0 !important;
      font-size: 1.05rem !important;
  }}

  /* Header e Footer banner */
  .apro-header {{
      width:100%; background-color:#2D5416; border-radius:10px;
      overflow:hidden; margin-bottom:1rem; box-shadow:0 3px 10px rgba(0,0,0,.3);
      display:flex; align-items:center; justify-content:center; padding:16px 0;
  }}
  .apro-header img {{ width:50%; max-width:480px; display:block; }}
  .apro-footer {{
      width:100%; background-color:#2D5416; border-radius:8px;
      overflow:hidden; margin-top:0.5rem; box-shadow:0 -2px 6px rgba(0,0,0,.15);
      display:flex; align-items:center; justify-content:center; padding:12px 0;
  }}
  .apro-footer img {{ width:38%; max-width:360px; display:block; }}
  .apro-caption {{
      text-align:center; color:#5a8c65; font-size:0.76rem; margin-top:5px;
  }}

  /* Simulador */
  .sim-header {{
      background: linear-gradient(145deg, #122518 0%, #1a3328 100%);
      border: 1px solid #243f2f; border-radius: 12px;
      padding: 14px 20px; margin-bottom: 14px;
      box-shadow: 0 4px 12px rgba(0,0,0,0.35);
  }}
  .sim-header h3 {{
      color: #e8f0eb !important;
      border-left: 4px solid #c8a415 !important;
      padding-left: 10px !important;
      margin: 0 !important;
      font-size: 1.05rem !important;
  }}
  .result-banner {{
      background: linear-gradient(145deg, #0d2918 0%, #122518 100%);
      border: 1px solid #2d5a3d;
      border-radius: 12px;
      padding: 18px 20px;
      margin-top: 14px;
      box-shadow: 0 4px 16px rgba(0,0,0,0.4);
  }}

  /* Form inputs no tema escuro */
  [data-testid="stNumberInput"] input,
  [data-testid="stSelectbox"] div {{
      background-color: #1a3025 !important;
      color: #e8f0eb !important;
      border-color: #2d5a3d !important;
  }}
  .stForm [data-testid="stFormSubmitButton"] button {{
      background: #c8a415 !important;
      color: #0e1a12 !important;
      font-weight: 700 !important;
      border: none !important;
      border-radius: 8px !important;
      font-size: 1rem !important;
  }}
  .stForm [data-testid="stFormSubmitButton"] button:hover {{
      background: #d4b020 !important;
  }}
</style>

<div class="apro-header">
  <img src="data:image/png;base64,{apro_b64}" />
</div>
""", unsafe_allow_html=True)

# ─── PALETA ───────────────────────────────────────────────────────────────────
_GOLD    = "#c8a415"
_GREEN   = "#4caf50"
_LIGHT   = "#e8f0eb"
_MUTED   = "#7aa88a"
_LABEL   = "#6a9978"
_BG_PLOT = "#111c16"
_BG_GRID = "#0f1f15"

CORES_CATS = [
    "#2E7D32", "#388E3C", "#558B2F", "#81C784",
    "#A5D6A7", "#c8a415", "#FFB300", "#FF8F00",
    "#E65100", "#BF360C", "#4E342E", "#6D4C41",
    "#1565C0", "#0288D1",
]

# ─── KPI HTML CARD ────────────────────────────────────────────────────────────
def _kpi(titulo, valor, sub="", cor="#c8a415", sub_cor="#4a6e55"):
    return (
        f'<div style="background:linear-gradient(145deg,#122518 0%,#1a3328 100%);'
        f'border:1px solid #243f2f;border-radius:16px;padding:18px 12px 14px;'
        f'text-align:center;min-height:108px;display:flex;flex-direction:column;'
        f'justify-content:center;box-shadow:0 6px 24px rgba(0,0,0,.45);">'
        f'<div style="color:#6a9978;font-size:.63rem;font-weight:700;'
        f'text-transform:uppercase;letter-spacing:.9px;margin-bottom:7px;'
        f'font-family:{_FONT}">{titulo}</div>'
        f'<div style="font-size:1.45rem;font-weight:800;line-height:1.1;'
        f'color:{cor};font-family:{_FONT}">{valor}</div>'
        f'<div style="color:{sub_cor};font-size:.62rem;margin-top:4px;'
        f'font-family:{_FONT}">{sub}</div>'
        f'</div>'
    )

# ─── DADOS DE CUSTO ──────────────────────────────────────────────────────────
# Produtividades médias MS (sc/ha)
PROD_SOJA   = 61.73
PROD_MILHO  = 84.0

# Preços de referência (fallback caso o arquivo não exista)
_PRECO_SOJA_FALLBACK  = 120.0
_PRECO_MILHO_FALLBACK =  54.0

SOJA_ATUAL = {
    "Safra": "2025/26",
    "Máquinas/Combustível":     180.90,
    "Sementes":                 589.20,
    "Tratamento de Semente":     38.50,
    "Corretivo de Solo":        289.80,
    "Fertilizante":           1_395.80,
    "Fungicida":                352.00,
    "Herbicida":                292.80,
    "Inseticida":               315.00,
    "Inoculantes/Adjuvante":     50.30,
    "Outros Variáveis":       1_670.34,
    "Juros":                    350.43,
    "Custo Fixo":               399.96,
    "Renda dos Fatores":        190.80,
}
SOJA_ANTERIOR = {
    "Safra": "2024/25",
    "Máquinas/Combustível":     179.10,
    "Sementes":                 568.80,
    "Tratamento de Semente":     42.00,
    "Corretivo de Solo":        418.50,
    "Fertilizante":           1_125.00,
    "Fungicida":                425.50,
    "Herbicida":                398.15,
    "Inseticida":               200.00,
    "Inoculantes/Adjuvante":     74.55,
    "Outros Variáveis":       1_641.00,
    "Juros":                    344.00,
    "Custo Fixo":               394.00,
    "Renda dos Fatores":        187.64,
}
MILHO_ATUAL = {
    "Safra": "2025/26",
    "Máquinas/Combustível":     213.00,
    "Sementes":                 820.00,
    "Tratamento de Semente":     20.21,
    "Semente de Cobertura":      45.00,
    "Corretivo de Solo":        230.00,
    "Fertilizante":           1_283.00,
    "Fungicida":                 88.25,
    "Herbicida":                 90.50,
    "Inseticida":               329.68,
    "Inoculantes/Adjuvante":     18.60,
    "Outros Variáveis":         710.24,
    "Juros":                    313.82,
    "Custo Fixo":               316.51,
    "Renda dos Fatores":        358.30,
}
MILHO_ANTERIOR = {
    "Safra": "2025",
    "Máquinas/Combustível":     200.00,
    "Sementes":                 750.00,
    "Tratamento de Semente":     18.00,
    "Semente de Cobertura":      42.00,
    "Corretivo de Solo":        220.00,
    "Fertilizante":           1_100.00,
    "Fungicida":                 88.00,
    "Herbicida":                 85.00,
    "Inseticida":               280.00,
    "Inoculantes/Adjuvante":     17.46,
    "Outros Variáveis":         652.18,
    "Juros":                    289.05,
    "Custo Fixo":               311.56,
    "Renda dos Fatores":        421.45,
}

SOJA_DEPREC    = 325.00
SOJA_IMPOSTOS  =  70.09
MILHO_DEPREC   = 250.00
MILHO_IMPOSTOS =  31.38

# Detalhamento de "Outros Variáveis" — fonte: planilhas Aprosoja/MS
SOJA_OUTROS_VARS = {
    "Seguro Agrícola":                  275.00,
    "Transporte Externo":               116.60,
    "Armazenagem":                      174.90,
    "Assistência Técnica":              109.54,
    "Impostos e Taxas":                  70.09,
    "Manutenção Máquinas/Implementos":  146.25,
    "Mão de Obra":                      147.18,
    "Despesas Administrativas":         630.77,
}
MILHO_OUTROS_VARS = {
    "Seguro Agrícola":                   37.36,
    "Transporte Externo":               159.75,
    "Armazenagem":                      109.84,
    "Assistência Técnica":               96.22,
    "Impostos e Taxas":                  31.38,
    "Manutenção Máquinas/Implementos":  112.50,
    "Mão de Obra":                       69.04,
    "Despesas Administrativas":          94.15,
}

HISTORICO_CUSTO = {
    "Safra":         ["2016/17","2017/18","2018/19","2019/20","2020/21","2021/22","2022/23","2023/24","2024/25","2025/26"],
    "Soja (R$/ha)":  [2_700,    3_000,    2_900,    3_050,    3_800,    5_100,    5_400,    5_450,    5_998,    6_116],
    "Milho (R$/ha)": [2_200,    2_300,    2_380,    2_500,    2_900,    3_900,    4_200,    4_000,    4_475,    4_837],
}

# ─── EXCEL COLETA ─────────────────────────────────────────────────────────────
EXCEL_DADOS = os.path.join(BASE_DIR, "Dados_Produtores.xlsx")
COLUNAS_EXCEL = [
    "Data","Município","Área (ha)","Produtividade (sc/ha)","Preço Venda (R$/sc)",
    "Sementes (R$)","Tratamento de Semente (R$)","Corretivo de Solo (R$)",
    "Fertilizante (R$)","Fungicida (R$)","Herbicida (R$)","Inseticida (R$)",
    "Inoculantes (R$)","Adjuvante (R$)","Op. c/ Máquinas (R$)","Seguro Agrícola (R$)",
    "Transporte Externo (R$)","Armazenagem (R$)","Assistência Técnica (R$)",
    "Manutenção Máquinas (R$)","Mão de Obra (R$)","Desp. Administrativas (R$)",
    "Juros/Financiamentos (R$)","Depreciação (R$)","Outros Custos (R$)",
    "CUSTO TOTAL (R$)","CUSTO/ha (R$)","CUSTO/sc (R$)",
    "RECEITA/ha (R$)","MARGEM/ha (R$)","PONTO DE EQUILÍBRIO (sc/ha)",
]

def salvar_produtor(municipio, valores, kpis):
    _pull_dados_github()
    if os.path.exists(EXCEL_DADOS):
        wb = openpyxl.load_workbook(EXCEL_DADOS)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    nome_aba = municipio[:31]
    if nome_aba not in wb.sheetnames:
        ws = wb.create_sheet(nome_aba)
        ws.append(COLUNAS_EXCEL)
        hf = PatternFill("solid", fgColor="1B4332")
        hft = Font(bold=True, color="FFFFFF")
        for ci in range(1, len(COLUNAS_EXCEL) + 1):
            c = ws.cell(row=1, column=ci)
            c.fill, c.font = hf, hft
            c.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
    else:
        ws = wb[nome_aba]

    ws.append([
        date.today().isoformat(), municipio,
        valores.get("area",0), valores.get("produtividade",0), valores.get("preco",0),
        valores.get("sementes",0), valores.get("trat_semente",0), valores.get("corretivo",0),
        valores.get("fertilizante",0), valores.get("fungicida",0), valores.get("herbicida",0),
        valores.get("inseticida",0), valores.get("inoculantes",0), valores.get("adjuvante",0),
        valores.get("maquinas",0), valores.get("seguro",0), valores.get("transporte",0),
        valores.get("armazenagem",0), valores.get("assist_tec",0), valores.get("manutencao",0),
        valores.get("mao_obra",0), valores.get("desp_admin",0),
        valores.get("juros",0), valores.get("deprec",0), valores.get("outros",0),
        kpis["custo_total_r"], kpis["custo_ha"], kpis["custo_sc"],
        kpis["receita_ha"], kpis["margem_ha"], kpis["ponto_eq"],
    ])
    wb.save(EXCEL_DADOS)


def _github_token() -> str:
    """Retorna o GITHUB_TOKEN a partir de variável de ambiente ou st.secrets."""
    token = os.environ.get("GITHUB_TOKEN", "")
    if not token:
        try:
            token = st.secrets.get("GITHUB_TOKEN", "")
        except Exception:
            pass
    return token or ""


def _pull_dados_github() -> bool:
    """Baixa Dados_Produtores.xlsx do GitHub para garantir dado mais recente antes de salvar."""
    import urllib.request, urllib.error, json as _json
    repo  = os.environ.get("GITHUB_REPO", "economia2-cyber/custodeproducao")
    token = _github_token()

    api     = f"https://api.github.com/repos/{repo}/contents/Dados_Produtores.xlsx"
    headers = {"Accept": "application/vnd.github.v3+json"}
    if token:
        headers["Authorization"] = f"token {token}"

    try:
        req = urllib.request.Request(api, headers=headers)
        with urllib.request.urlopen(req) as r:
            data = _json.loads(r.read())
        content = base64.b64decode(data["content"].replace("\n", ""))
        with open(EXCEL_DADOS, "wb") as f:
            f.write(content)
        return True
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return True  # arquivo ainda não existe no repositório
        return False
    except Exception:
        return False


def _push_dados_github() -> tuple:
    """Commita Dados_Produtores.xlsx de volta ao repositório via GitHub API."""
    import urllib.request, urllib.error, json as _json
    token = _github_token()
    repo  = os.environ.get("GITHUB_REPO", "economia2-cyber/custodeproducao")
    if not token:
        return False, "GITHUB_TOKEN não configurado nas variáveis de ambiente"
    if not os.path.exists(EXCEL_DADOS):
        return False, "Arquivo local não encontrado"

    api = f"https://api.github.com/repos/{repo}/contents/Dados_Produtores.xlsx"
    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json",
        "Content-Type": "application/json",
    }

    sha = None
    try:
        req = urllib.request.Request(api, headers=headers)
        with urllib.request.urlopen(req) as r:
            sha = _json.loads(r.read()).get("sha")
    except urllib.error.HTTPError as e:
        if e.code != 404:
            return False, f"GitHub API {e.code}"
    except Exception as e:
        return False, str(e)

    with open(EXCEL_DADOS, "rb") as f:
        content_b64 = base64.b64encode(f.read()).decode()

    payload: dict = {
        "message": f"Dados produtor — {date.today().isoformat()}",
        "content": content_b64,
    }
    if sha:
        payload["sha"] = sha

    try:
        body = _json.dumps(payload).encode()
        req = urllib.request.Request(api, data=body, headers=headers, method="PUT")
        with urllib.request.urlopen(req):
            pass
        return True, ""
    except Exception as e:
        return False, str(e)


def gerar_excel_composicao(cultura_label, safra_label, cats, valores, total, logo_path, outros_vars=None):
    """Gera Excel de composição do custo com logo visível, detalhamento de Outros Variáveis e bloqueio."""
    wb = openpyxl.Workbook()
    ws = wb.active
    _titulo_aba = f"Custo {cultura_label} {safra_label}".replace("/", "-").replace("\\", "-")[:31]
    ws.title = _titulo_aba

    # Paleta
    _VD  = "2D5416"   # verde escuro (cabeçalho principal)
    _VM  = "4a7a28"   # verde médio (seção Outros)
    _VE  = "1B4332"   # verde bem escuro (sub-cabeçalhos)
    _ALT = "EAF4DF"   # verde clarinho (linhas alternadas)
    _BCO = "FFFFFF"

    def _borda():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    # ── Linha 1: fundo verde escuro para logo branca ficar visível ────────────
    ws.row_dimensions[1].height = 58
    for col in range(1, 5):
        ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor=_VD)

    if os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.width  = 260
        img.height = 58
        ws.add_image(img, "A1")

    ws.row_dimensions[2].height = 5  # espaçador

    # ── Linha 3: Título ────────────────────────────────────────────────────────
    ws.merge_cells("A3:D3")
    c = ws["A3"]
    c.value     = f"Composição do Custo de Produção — {cultura_label} | {safra_label}"
    c.font      = Font(bold=True, size=13, color=_BCO)
    c.fill      = PatternFill("solid", fgColor=_VD)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 24

    # ── Linha 4: Fonte / Data ──────────────────────────────────────────────────
    ws.merge_cells("A4:D4")
    c = ws["A4"]
    c.value     = f"Fonte: Aprosoja/MS   |   Extração: {date.today().strftime('%d/%m/%Y')}"
    c.font      = Font(italic=True, size=9, color="5A8C65")
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[4].height = 14

    ws.row_dimensions[5].height = 5  # espaçador

    # ── Linha 6: Cabeçalhos da tabela principal ────────────────────────────────
    for col, h in enumerate(["Componente de Custo", "Valor (R$/ha)", "Participação (%)"], 1):
        c = ws.cell(row=6, column=col, value=h)
        c.font      = Font(bold=True, size=10, color=_BCO)
        c.fill      = PatternFill("solid", fgColor=_VE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _borda()
    ws.row_dimensions[6].height = 17

    # ── Linhas de dados (composição principal) ─────────────────────────────────
    for i, (cat, val) in enumerate(zip(cats, valores)):
        row  = 7 + i
        pct  = val / total if total else 0
        fill = PatternFill("solid", fgColor=_ALT) if i % 2 == 0 else PatternFill("solid", fgColor=_BCO)
        c1 = ws.cell(row=row, column=1, value=cat)
        c2 = ws.cell(row=row, column=2, value=val)
        c3 = ws.cell(row=row, column=3, value=pct)
        for cx in (c1, c2, c3):
            cx.font   = Font(size=10)
            cx.fill   = fill
            cx.border = _borda()
        c2.number_format = '"R$"\\ #,##0.00'
        c2.alignment     = Alignment(horizontal="right")
        c3.number_format = "0.00%"
        c3.alignment     = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 15

    # ── Linha de Total ─────────────────────────────────────────────────────────
    row_total = 7 + len(cats)
    for col, (v, fmt, align) in enumerate(
        [("TOTAL", None, "left"), (total, '"R$"\\ #,##0.00', "right"), (1.0, "0.00%", "center")], 1
    ):
        c = ws.cell(row=row_total, column=col, value=v)
        c.font      = Font(bold=True, size=10, color=_BCO)
        c.fill      = PatternFill("solid", fgColor=_VD)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border    = _borda()
        if fmt:
            c.number_format = fmt
    ws.row_dimensions[row_total].height = 17

    # ── Seção: Detalhamento — Outros Variáveis ─────────────────────────────────
    if outros_vars:
        r = row_total + 2
        ws.row_dimensions[r - 1].height = 10  # espaçador

        ws.merge_cells(f"A{r}:D{r}")
        c = ws.cell(row=r, column=1, value="Detalhamento — Outros Variáveis")
        c.font      = Font(bold=True, size=11, color=_BCO)
        c.fill      = PatternFill("solid", fgColor=_VM)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _borda()
        ws.row_dimensions[r].height = 20
        r += 1

        for col, h in enumerate(
            ["Item", "Valor (R$/ha)", "% s/ Outros Variáveis", "% s/ Custo Total"], 1
        ):
            c = ws.cell(row=r, column=col, value=h)
            c.font      = Font(bold=True, size=9, color=_BCO)
            c.fill      = PatternFill("solid", fgColor=_VE)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = _borda()
        ws.row_dimensions[r].height = 15
        r += 1

        total_outros = sum(outros_vars.values())
        for i, (item, val) in enumerate(outros_vars.items()):
            pct_ov = val / total_outros if total_outros else 0
            pct_ct = val / total        if total        else 0
            fill   = PatternFill("solid", fgColor=_ALT) if i % 2 == 0 else PatternFill("solid", fgColor=_BCO)
            for col, (v, fmt, align) in enumerate(
                [(item, None, "left"), (val, '"R$"\\ #,##0.00', "right"),
                 (pct_ov, "0.00%", "center"), (pct_ct, "0.00%", "center")], 1
            ):
                c = ws.cell(row=r, column=col, value=v)
                c.font      = Font(size=9)
                c.fill      = fill
                c.border    = _borda()
                c.alignment = Alignment(horizontal=align)
                if fmt:
                    c.number_format = fmt
            ws.row_dimensions[r].height = 14
            r += 1

        # Subtotal Outros Variáveis
        pct_ov_total = total_outros / total if total else 0
        for col, (v, fmt, align) in enumerate(
            [("SUBTOTAL", None, "left"), (total_outros, '"R$"\\ #,##0.00', "right"),
             (1.0, "0.00%", "center"), (pct_ov_total, "0.00%", "center")], 1
        ):
            c = ws.cell(row=r, column=col, value=v)
            c.font      = Font(bold=True, size=9, color=_BCO)
            c.fill      = PatternFill("solid", fgColor=_VM)
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border    = _borda()
            if fmt:
                c.number_format = fmt
        ws.row_dimensions[r].height = 16

    # ── Larguras de coluna ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18

    # ── Proteção (objetos/imagens ficam visíveis) ──────────────────────────────
    ws.protection.sheet    = True
    ws.protection.password = "aprosoja"
    ws.protection.objects  = False

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─── CARREGAMENTO ─────────────────────────────────────────────────────────────
_EXCLUIR_LINHAS = {"resultado ponderado", "resultados aritmético", "resultados aritmetico"}

def _is_municipio(nome):
    return isinstance(nome, str) and nome.strip() and nome.strip().lower() not in _EXCLUIR_LINHAS


@st.cache_data
def load_produtividade_soja():
    path = os.path.join(BASE_DIR, "Produtividade_Soja.xlsx")
    if not os.path.exists(path):
        return pd.DataFrame(columns=["Município", "Produtividade (sc/ha)"])
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Principal"]
    rows = [(r[0].strip(), float(r[1])) for r in ws.iter_rows(min_row=4, values_only=True)
            if _is_municipio(r[0]) and isinstance(r[1], (int, float)) and r[1] > 0]
    wb.close()
    return (pd.DataFrame(rows, columns=["Município", "Produtividade (sc/ha)"])
              .drop_duplicates(subset="Município", keep="first")
              .reset_index(drop=True))


@st.cache_data
def load_produtividade_milho():
    path = os.path.join(BASE_DIR, "Produtividade_Milho.xlsx")
    if not os.path.exists(path):
        return pd.DataFrame(columns=["Município", "Produtividade (sc/ha)"])
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Principal"]
    rows = [(r[0].strip(), float(r[1])) for r in ws.iter_rows(min_row=4, values_only=True)
            if _is_municipio(r[0]) and isinstance(r[1], (int, float)) and r[1] > 0]
    wb.close()
    return (pd.DataFrame(rows, columns=["Município", "Produtividade (sc/ha)"])
              .drop_duplicates(subset="Município", keep="first")
              .reset_index(drop=True))


@st.cache_data
def load_prod_media_soja():
    path = os.path.join(BASE_DIR, "Produtividade_Soja.xlsx")
    if not os.path.exists(path):
        return PROD_SOJA
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Principal"]
    for row in ws.iter_rows(min_row=4, values_only=True):
        if isinstance(row[0], str) and "resultado ponderado" in row[0].strip().lower():
            wb.close()
            return float(row[1]) if isinstance(row[1], (int, float)) else PROD_SOJA
    wb.close()
    return PROD_SOJA


@st.cache_data
def load_prod_media_milho():
    path = os.path.join(BASE_DIR, "Produtividade_Milho.xlsx")
    if not os.path.exists(path):
        return PROD_MILHO
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Principal"]
    for row in ws.iter_rows(min_row=4, values_only=True):
        if isinstance(row[0], str) and "resultado ponderado" in row[0].strip().lower():
            wb.close()
            return float(row[1]) if isinstance(row[1], (int, float)) else PROD_MILHO
    wb.close()
    return PROD_MILHO


@st.cache_data(ttl=300)
def load_precos():
    """
    Lê Preços 2.0.xlsx e retorna a média diária mais recente de soja e milho.
    Retorna (preco_soja, preco_milho, data_str).
    """
    path = os.path.join(BASE_DIR, "Preços 2.0.xlsx")
    if not os.path.exists(path):
        return _PRECO_SOJA_FALLBACK, _PRECO_MILHO_FALLBACK, "referência"

    # Colunas de disponível por fonte
    SOJA_COLS  = [2, 6, 8, 10, 12, 14, 16, 18]
    MILHO_COLS = [4, 7, 9, 11, 13, 15, 17, 19]

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    records = []
    current_date = None
    for row in ws.iter_rows(min_row=3, values_only=True):
        d = row[0]
        if hasattr(d, "date"):          # datetime
            current_date = d
        elif isinstance(d, date):       # date puro
            current_date = d
        if current_date is None:
            continue
        sv = [row[c] for c in SOJA_COLS  if c < len(row) and isinstance(row[c], (int, float))]
        mv = [row[c] for c in MILHO_COLS if c < len(row) and isinstance(row[c], (int, float))]
        if sv or mv:
            records.append({
                "data":        current_date,
                "media_soja":  sum(sv) / len(sv) if sv else None,
                "media_milho": sum(mv) / len(mv) if mv else None,
            })
    wb.close()

    if not records:
        return _PRECO_SOJA_FALLBACK, _PRECO_MILHO_FALLBACK, "referência"

    df = pd.DataFrame(records)
    df["data"] = pd.to_datetime(df["data"])
    daily = (
        df.groupby("data")
          .agg(media_soja=("media_soja", "mean"), media_milho=("media_milho", "mean"))
          .reset_index()
          .dropna(subset=["media_soja", "media_milho"])
          .sort_values("data", ascending=False)
    )

    if daily.empty:
        return _PRECO_SOJA_FALLBACK, _PRECO_MILHO_FALLBACK, "referência"

    ultima = daily.iloc[0]
    data_str = ultima["data"].strftime("%d/%m/%Y")
    return round(float(ultima["media_soja"]), 2), round(float(ultima["media_milho"]), 2), data_str


@st.cache_data
def load_regioes():
    """Lê Regiões.xlsx → (dict regiao→[munis], dict muni→regiao)."""
    path = os.path.join(BASE_DIR, "Regiões.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Planilha1"]
    reg_munis: dict = {}
    muni_reg:  dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        reg, mun = row[0], row[1]
        if reg and mun:
            reg, mun = str(reg).strip(), str(mun).strip()
            reg_munis.setdefault(reg, []).append(mun)
            muni_reg[mun] = reg
    wb.close()
    return reg_munis, muni_reg


# ─── CÁLCULOS ─────────────────────────────────────────────────────────────────
def custo_total(d):
    return sum(v for k, v in d.items() if k != "Safra")


def fmt_brl(v, dec=2):
    s = f"{v:,.{dec}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"


def calcular_kpis(cultura, preco_soja, preco_milho):
    if cultura == "Soja IPRO":
        custo    = custo_total(SOJA_ATUAL)
        preco    = preco_soja
        prod     = load_prod_media_soja()
        deprec   = SOJA_DEPREC
        juros    = SOJA_ATUAL["Juros"]
        impostos = SOJA_IMPOSTOS
        custo_ant = custo_total(SOJA_ANTERIOR)
    else:
        custo    = custo_total(MILHO_ATUAL)
        preco    = preco_milho
        prod     = load_prod_media_milho()
        deprec   = MILHO_DEPREC
        juros    = MILHO_ATUAL["Juros"]
        impostos = MILHO_IMPOSTOS
        custo_ant = custo_total(MILHO_ANTERIOR)

    receita  = prod * preco
    margem   = receita - custo
    lajida   = margem + juros + deprec + impostos
    ponto_eq = custo / preco if preco else 0
    custo_sc = custo / prod  if prod  else 0
    var      = (custo - custo_ant) / custo_ant * 100 if custo_ant else 0
    return dict(custo_ha=custo, custo_sc=custo_sc, receita_ha=receita,
                margem_ha=margem, lajida=lajida, ponto_eq=ponto_eq,
                preco=preco, prod=prod, var_custo=var, custo_ant=custo_ant)


# ─── SIDEBAR ─────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(
        f'<img src="data:image/png;base64,{apro_b64}" '
        f'style="width:100%;border-radius:8px;margin-bottom:12px;">',
        unsafe_allow_html=True,
    )
    st.markdown("---")

    # Botões Soja / Milho
    _sel_now = st.session_state.get("_cultura_sel", "Soja IPRO")
    _c1, _c2 = st.columns(2)
    with _c1:
        if st.button("🌱 SOJA", use_container_width=True, key="btn_soja",
                     type="primary" if _sel_now == "Soja IPRO" else "secondary"):
            st.session_state["_cultura_sel"] = "Soja IPRO"
            st.rerun()
    with _c2:
        if st.button("🌽 MILHO", use_container_width=True, key="btn_milho",
                     type="primary" if _sel_now == "Milho" else "secondary"):
            st.session_state["_cultura_sel"] = "Milho"
            st.rerun()

    # JS: contorno vermelho apenas no botão selecionado (igual ao Painel Exemplo)
    _lbl_sel = "SOJA" if _sel_now == "Soja IPRO" else "MILHO"
    components.html(f"""
<script>
(function(){{
  var SEL = "{_lbl_sel}";
  function applyStyles(){{
    var sb = window.parent.document.querySelector('[data-testid="stSidebar"]');
    if(!sb) return false;
    var hb = sb.querySelector('[data-testid="stHorizontalBlock"]');
    if(!hb) return false;
    var btns = hb.querySelectorAll('button');
    if(!btns.length) return false;
    hb.style.setProperty('padding','8px 4px','important');
    btns.forEach(function(b){{
      var isSel = b.textContent.indexOf(SEL) !== -1;
      b.style.setProperty('background','#3a6b1a','important');
      b.style.setProperty('background-color','#3a6b1a','important');
      b.style.setProperty('border-radius','14px','important');
      b.style.setProperty('font-weight', isSel?'700':'600','important');
      b.style.setProperty('color', isSel?'#ffffff':'#d4edaa','important');
      b.style.setProperty('border', isSel?'3px solid #e74c3c':'2px solid #5a9b2a','important');
      b.style.setProperty('box-shadow', isSel?'0 0 10px rgba(231,76,60,0.4)':'none','important');
    }});
    return true;
  }}
  if(!applyStyles()){{
    var t=0,iv=setInterval(function(){{if(applyStyles()||++t>20)clearInterval(iv);}},100);
  }}
}})();
</script>
""", height=0)

    cultura_sel = _sel_now
    st.markdown("---")

    # Carrega dados de produtividade e regiões
    df_prod    = load_produtividade_soja() if cultura_sel == "Soja IPRO" else load_produtividade_milho()
    reg_munis, muni_reg = load_regioes()
    todas_regioes = sorted(reg_munis.keys())

    regioes_sel = st.multiselect(
        "📍 Selecione sua região",
        options=todas_regioes,
        placeholder="Todas as regiões",
    )

    st.markdown("---")

    # Botão de aba "Custo" — visual estático elegante
    st.markdown(
        '<div style="background:linear-gradient(135deg,#1a4a28,#2d6b3a);'
        'border:2px solid #c8a415;border-radius:12px;padding:10px 14px;'
        'text-align:center;cursor:default;box-shadow:0 0 10px rgba(200,164,21,0.25);">'
        '<span style="color:#c8a415;font-weight:700;font-size:0.9rem;'
        'letter-spacing:1px;text-transform:uppercase;">💰 Custo de Produção</span>'
        '</div>',
        unsafe_allow_html=True,
    )

    st.markdown("---")
    st.caption("Aprosoja/MS · Safra 2025/2026")


# ─── DADOS FILTRADOS ──────────────────────────────────────────────────────────
# Municípios das regiões selecionadas
if regioes_sel:
    munis_filtro = [m for r in regioes_sel for m in reg_munis.get(r, [])]
    df_munic = df_prod[df_prod["Município"].isin(munis_filtro)].copy()
else:
    df_munic = df_prod.copy()

# Adiciona coluna Região usando mapeamento da planilha
df_munic = df_munic.copy()
df_munic["Região"] = df_munic["Município"].map(muni_reg).fillna("Outras")

preco_soja_atual, preco_milho_atual, data_preco = load_precos()
kpis = calcular_kpis(cultura_sel, preco_soja_atual, preco_milho_atual)

if cultura_sel == "Soja IPRO":
    bd_atual  = SOJA_ATUAL
    bd_ant    = SOJA_ANTERIOR
    lbl_at    = "Safra 25/26"
    lbl_ant   = "Safra 24/25"
else:
    bd_atual  = MILHO_ATUAL
    bd_ant    = MILHO_ANTERIOR
    lbl_at    = "Safra 25/26"
    lbl_ant   = "Safra 2025"


# ─── CONTEÚDO ─────────────────────────────────────────────────────────────────

# Título
_cultura_display = cultura_sel.replace(" IPRO", "")
st.markdown(
    f'<h1 style="color:{_LIGHT};font-size:1.85rem;font-weight:800;'
    f'letter-spacing:2px;margin:0 0 2px 0;line-height:1.2;font-family:{_FONT}">'
    f'CUSTO DE PRODUÇÃO <span style="color:{_GOLD}">{_cultura_display.upper()}</span>'
    f' <span style="font-size:1.2rem;font-weight:400;color:{_MUTED}">| Mato Grosso do Sul</span>'
    f'</h1>',
    unsafe_allow_html=True,
)
st.caption("Safra 2025/2026 · Fonte: Aprosoja/MS")
st.divider()

# ── KPIs ─────────────────────────────────────────────────────────────────────
# Custo vs período anterior: aumento = ruim = vermelho; queda = bom = verde
_var = kpis["var_custo"]
var_cor  = "#e07878" if _var > 0 else (_GREEN if _var < 0 else _MUTED)
var_str  = f"{'▲' if _var > 0 else '▼' if _var < 0 else '–'} {abs(_var):.1f}% vs safra anterior"

marg_cor   = _GREEN if kpis["margem_ha"] >= 0 else "#e07878"
lajida_cor = _GREEN if kpis["lajida"]    >= 0 else "#e07878"

# Ponto de equilíbrio: alerta se produtividade abaixo do ponto
pe_cor = "#e07878" if kpis["prod"] < kpis["ponto_eq"] else _GREEN

rec_sub = f"@ {fmt_brl(kpis['preco'])}/sc · {data_preco}"

c1, c2, c3, c4, c5, c6 = st.columns(6)
with c1: st.markdown(_kpi("Custo Total / ha",     fmt_brl(kpis["custo_ha"]),        var_str,  _GOLD,  var_cor),   unsafe_allow_html=True)
with c2: st.markdown(_kpi("Custo / Saca",         fmt_brl(kpis["custo_sc"]),        "R$/sc (60 kg)", _LIGHT),     unsafe_allow_html=True)
with c3: st.markdown(_kpi("Receita / ha",         fmt_brl(kpis["receita_ha"]),      rec_sub,  _LIGHT),            unsafe_allow_html=True)
with c4: st.markdown(_kpi("Margem Líquida / ha",  fmt_brl(kpis["margem_ha"]),       "",       marg_cor),          unsafe_allow_html=True)
with c5: st.markdown(_kpi("LAJIDA / ha",          fmt_brl(kpis["lajida"]),          "",       lajida_cor),        unsafe_allow_html=True)
with c6: st.markdown(_kpi("Ponto de Equilíbrio",  f"{kpis['ponto_eq']:.1f} sc/ha", f"Prod. média: {kpis['prod']:.1f} sc/ha", pe_cor), unsafe_allow_html=True)

st.divider()

# ── Tabela municípios ─────────────────────────────────────────────────────────
st.markdown('<div class="card-titulo"><h3>📊 Custo de Produção por Município — Mato Grosso do Sul</h3></div>', unsafe_allow_html=True)

df_tab = df_munic.copy()
df_tab["Custo/sc (R$)"]   = (kpis["custo_ha"] / df_tab["Produtividade (sc/ha)"]).round(2)
df_tab["Receita/ha (R$)"] = (df_tab["Produtividade (sc/ha)"] * kpis["preco"]).round(2)
df_tab["Margem/ha (R$)"]  = (df_tab["Receita/ha (R$)"] - kpis["custo_ha"]).round(2)

df_tab = df_tab[[
    "Município", "Produtividade (sc/ha)", "Custo/sc (R$)", "Receita/ha (R$)", "Margem/ha (R$)",
]].sort_values("Município", ascending=True)

def _cor_marg(v):
    if isinstance(v, (int, float)):
        return f"color:{'#4caf50' if v >= 0 else '#e07878'};font-weight:700"
    return ""

styled = (
    df_tab.style
    .format({
        "Produtividade (sc/ha)": "{:.2f}",
        "Custo/sc (R$)":         "R$ {:,.2f}",
        "Receita/ha (R$)":       "R$ {:,.2f}",
        "Margem/ha (R$)":        "R$ {:,.2f}",
    })
    .map(_cor_marg, subset=["Margem/ha (R$)"])
    .set_properties(**{"background-color": "#122518", "color": "#e8f0eb"})
    .set_properties(**{"text-align": "right"}, subset=[
        "Produtividade (sc/ha)", "Custo/sc (R$)", "Receita/ha (R$)", "Margem/ha (R$)",
    ])
)
st.dataframe(styled, use_container_width=True, height=360, hide_index=True)
st.divider()

# ── Evolução CONAB ────────────────────────────────────────────────────────────
st.markdown('<div class="card-titulo"><h3>📈 Evolução do Custo de Produção — Mato Grosso do Sul</h3></div>', unsafe_allow_html=True)
st.caption("Fonte: CONAB (histórico) e Aprosoja/MS (safras 2024/25 e 2025/26)")

df_hist = pd.DataFrame(HISTORICO_CUSTO)
n = len(df_hist)

if cultura_sel == "Soja IPRO":
    _hist_col, _hist_cor = "Soja (R$/ha)", _GREEN
    _hist_font_cor = "white"
else:
    _hist_col, _hist_cor = "Milho (R$/ha)", _GOLD
    _hist_font_cor = "#0e1a12"

fig_hist = go.Figure()
fig_hist.add_trace(go.Scatter(
    x=df_hist["Safra"], y=df_hist[_hist_col],
    name=_hist_col.split(" ")[0],
    line=dict(color=_hist_cor, width=2.5),
    mode="lines+markers",
    marker=dict(size=[7]*(n-1)+[13], color=_hist_cor, line=dict(width=[0]*(n-1)+[2.5], color="white")),
    hovertemplate=f"<b>{_hist_col.split()[0]} %{{x}}</b>: R$ %{{y:,.0f}}/ha<extra></extra>",
))
fig_hist.add_annotation(
    x=df_hist["Safra"].iloc[-1], y=df_hist[_hist_col].iloc[-1],
    text=f"<b>R$ {df_hist[_hist_col].iloc[-1]:,.0f}</b>",
    showarrow=False, yshift=18,
    bgcolor=_hist_cor, bordercolor=_hist_cor, borderwidth=2, borderpad=5,
    font=dict(size=11, color=_hist_font_cor),
)

fig_hist.update_layout(
    height=370,
    paper_bgcolor=_BG_PLOT, plot_bgcolor=_BG_GRID,
    xaxis=dict(showgrid=False, zeroline=False, tickfont=dict(family=_FONT, size=11, color=_MUTED)),
    yaxis=dict(showgrid=True, gridcolor="rgba(76,175,80,0.12)", zeroline=False,
               tickprefix="R$ ", tickfont=dict(family=_FONT, size=10, color=_MUTED)),
    showlegend=False,
    hovermode="x unified",
    margin=dict(t=40, b=20, l=10, r=10),
    font=dict(family=_FONT, color=_LIGHT),
)
_add_watermark(fig_hist)
st.plotly_chart(fig_hist, use_container_width=True)

st.divider()

# ── Rosca + Barras ────────────────────────────────────────────────────────────
col_donut, col_bar = st.columns([1, 1.6])

cats   = [k for k in bd_atual if k != "Safra"]
v_at   = [bd_atual[k]       for k in cats]
v_ant  = [bd_ant.get(k, 0)  for k in cats]
total_at  = sum(v_at)
total_ant = sum(v_ant)

_outros_vars_detalhe = SOJA_OUTROS_VARS if cultura_sel == "Soja IPRO" else MILHO_OUTROS_VARS
_excel_bytes = gerar_excel_composicao(
    cultura_label = _cultura_display,
    safra_label   = lbl_at,
    cats          = cats,
    valores       = v_at,
    total         = total_at,
    logo_path     = _logo_path,
    outros_vars   = _outros_vars_detalhe,
)

with col_donut:
    _icone_donut = "🌱" if cultura_sel == "Soja IPRO" else "🌽"
    _xl_b64  = base64.b64encode(_excel_bytes).decode()
    _fn_xlsx = f"Composicao_Custo_{_cultura_display.replace(' ','_')}_{lbl_at.replace('/','_')}.xlsx"
    _dl_link = (
        f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;'
        f'base64,{_xl_b64}" download="{_fn_xlsx}" title="Exportar para Excel" '
        f'style="color:{_GOLD};font-size:1.05rem;text-decoration:none;flex-shrink:0;'
        f'background:rgba(200,164,21,0.13);border:1px solid rgba(200,164,21,0.35);'
        f'border-radius:6px;padding:2px 9px;margin-left:auto;">📥</a>'
    )
    st.markdown(
        f'<div class="card-titulo" style="display:flex;align-items:center;gap:8px;">'
        f'<h3 style="margin:0 !important;flex:1;">'
        f'{_icone_donut} Composição do Custo — {lbl_at}</h3>'
        f'{_dl_link}</div>',
        unsafe_allow_html=True,
    )

    fig_donut = go.Figure(go.Pie(
        labels=cats, values=v_at, hole=0.50,
        marker_colors=CORES_CATS[:len(cats)],
        textinfo="label+percent",
        textfont=dict(family=_FONT, size=8.5, color=_LIGHT),
        outsidetextfont=dict(family=_FONT, size=8.5, color=_LIGHT),
        hovertemplate="<b>%{label}</b><br>R$ %{value:,.2f}/ha<br>%{percent}<extra></extra>",
        automargin=True,
    ))
    fig_donut.add_annotation(
        text=f"<b>{fmt_brl(total_at)}</b><br><span style='font-size:10px;color:{_MUTED}'>Total/ha</span>",
        x=0.5, y=0.5, showarrow=False,
        font=dict(size=12, color=_LIGHT, family=_FONT),
    )
    fig_donut.update_layout(
        height=520, showlegend=False,
        paper_bgcolor=_BG_PLOT,
        margin=dict(t=40, b=40, l=60, r=60),
        font=dict(family=_FONT),
    )
    _add_watermark(fig_donut)
    st.plotly_chart(fig_donut, use_container_width=True)

with col_bar:
    st.markdown(f'<div class="card-titulo"><h3>📊 Composição do Custo — {lbl_ant} vs {lbl_at}</h3></div>', unsafe_allow_html=True)

    fig_bar = go.Figure()
    for i, (cat, cor) in enumerate(zip(cats, CORES_CATS[:len(cats)])):
        fig_bar.add_trace(go.Bar(
            name=cat,
            x=[lbl_ant, lbl_at],
            y=[v_ant[i], v_at[i]],
            marker_color=cor,
            width=0.35,
            hovertemplate=f"<b>{cat}</b><br>%{{x}}: R$ %{{y:,.2f}}/ha<extra></extra>",
        ))

    for x_pos, total_val in [(lbl_ant, total_ant), (lbl_at, total_at)]:
        fig_bar.add_annotation(
            x=x_pos, y=total_val,
            text=f"<b>{fmt_brl(total_val)}</b>",
            showarrow=False, yshift=14,
            bgcolor=_BG_PLOT, bordercolor=_GOLD, borderwidth=2, borderpad=4,
            font=dict(size=11, color=_GOLD, family=_FONT),
        )

    _y_max = max(total_ant, total_at) * 1.18   # margem para anotação de total
    fig_bar.update_layout(
        barmode="stack", height=420,
        bargap=0.65,                            # espaço entre as barras (0=sem espaço, 1=bar invisível)
        paper_bgcolor=_BG_PLOT, plot_bgcolor=_BG_GRID,
        xaxis=dict(showgrid=False, zeroline=False, tickfont=dict(family=_FONT, size=13, color=_LIGHT)),
        yaxis=dict(title="R$/ha", showgrid=True, gridcolor="rgba(76,175,80,0.12)",
                   zeroline=False, tickprefix="R$ ", tickfont=dict(family=_FONT, size=10, color=_MUTED),
                   range=[0, _y_max]),
        legend=dict(orientation="v", yanchor="middle", y=0.5, xanchor="left", x=1.01,
                    font=dict(family=_FONT, size=9, color=_LIGHT), bgcolor="rgba(0,0,0,0)"),
        margin=dict(t=10, b=20, l=10, r=165),
        font=dict(family=_FONT, color=_LIGHT),
    )
    _add_watermark(fig_bar)
    st.plotly_chart(fig_bar, use_container_width=True)

    # ── Nota: Outros Variáveis ────────────────────────────────────────────────
    st.markdown(
        f'<div style="background:linear-gradient(145deg,#0d2010 0%,#112218 100%);'
        f'border:1px solid #1e4a2a;border-left:3px solid {_GOLD};border-radius:8px;'
        f'padding:5px 10px;margin-top:4px;">'
        f'<span style="color:{_GOLD};font-weight:700;font-size:0.60rem;'
        f'font-family:{_FONT};text-transform:uppercase;letter-spacing:0.6px;">'
        f'ℹ️ Nota — Outros Variáveis</span>'
        f'<p style="color:{_MUTED};font-size:0.60rem;font-family:{_FONT};margin:3px 0 0;line-height:1.35;">'
        f'O item <b style="color:{_LIGHT}">Outros Variáveis</b> agrega os demais custos variáveis não detalhados '
        f'individualmente, compreendendo: '
        f'<b style="color:{_LIGHT}">Mão de Obra</b>; '
        f'<b style="color:{_LIGHT}">Transporte Externo</b> (frete e sacaria); '
        f'<b style="color:{_LIGHT}">Armazenagem</b> (recepção, secagem e estocagem); '
        f'<b style="color:{_LIGHT}">Seguro Agrícola</b> (PROAGRO Mais / apólice privada); '
        f'<b style="color:{_LIGHT}">Assistência Técnica</b>; '
        f'<b style="color:{_LIGHT}">Classificação e Amostragem</b>; '
        f'<b style="color:{_LIGHT}">FUNRURAL</b> (2,5% s/ receita bruta) '
        f'e <b style="color:{_LIGHT}">Despesas Administrativas</b>.'
        f'</p></div>',
        unsafe_allow_html=True,
    )

st.divider()

# ── Fontes ────────────────────────────────────────────────────────────────────
st.markdown(
    f'<p style="font-size:0.76rem;color:#5a8c65;font-style:italic;'
    f'font-family:{_FONT};text-align:center;">'
    f'📌 <b style="color:{_MUTED}">Fontes:</b> '
    f'Aprosoja/MS — Custo de Produção Soja e Milho, Safra 2025/2026 · '
    f'Aprosoja/MS — Estimativa de Produtividade por Município · '
    f'CONAB — Custo de Produção (séries históricas)</p>',
    unsafe_allow_html=True,
)


# ─── SIMULADOR ───────────────────────────────────────────────────────────────
st.divider()
st.markdown('<div class="sim-header"><h3>🌱 Simulador de Custo — Informe os Dados da Sua Lavoura</h3></div>', unsafe_allow_html=True)
st.markdown(
    f'<p style="color:{_MUTED};font-size:0.85rem;font-family:{_FONT};margin-bottom:12px">'
    f'Preencha os valores gastos e simule os indicadores da sua lavoura. '
    f'<b style="color:{_LIGHT}">Campos obrigatórios:</b> Município, Área e Produtividade.</p>',
    unsafe_allow_html=True,
)

if "sim_clear_v" not in st.session_state:
    st.session_state["sim_clear_v"] = 0

_preco_default = preco_soja_atual if cultura_sel == "Soja IPRO" else preco_milho_atual
_deprec_ref    = SOJA_DEPREC      if cultura_sel == "Soja IPRO" else MILHO_DEPREC

with st.form(f"sim_form_{st.session_state['sim_clear_v']}"):
    r1c1, r1c2, r1c3, r1c4 = st.columns(4)
    with r1c1:
        _todos_sim = sorted(set(df_prod["Município"].tolist()) | set(muni_reg.keys()))
        sim_mun  = st.selectbox("Município *", [""] + _todos_sim)
    with r1c2:
        sim_area = st.number_input("Área (ha) *", min_value=0.0, value=0.0, step=10.0, format="%.1f")
    with r1c3:
        sim_prod = st.number_input("Produtividade (sc/ha) *", min_value=0.0, value=0.0, step=1.0, format="%.1f")
    with r1c4:
        sim_preco = st.number_input("Preço de venda (R$/sc)", min_value=0.0, value=float(_preco_default), step=1.0, format="%.2f")

    st.markdown(f'<p style="color:{_MUTED};font-size:0.82rem;font-family:{_FONT};margin:8px 0 4px">Informe o valor total gasto em cada item (R$):</p>', unsafe_allow_html=True)

    cA, cB, cC = st.columns(3)
    with cA:
        st.markdown(f'<small style="color:{_LABEL};font-family:{_FONT}">INSUMOS DA LAVOURA</small>', unsafe_allow_html=True)
        v_sem  = st.number_input("Sementes",             0.0, step=100.0, format="%.2f")
        v_trat = st.number_input("Tratamento de Semente",0.0, step=50.0,  format="%.2f")
        v_corr = st.number_input("Corretivo de Solo",    0.0, step=100.0, format="%.2f")
        v_fert = st.number_input("Fertilizante",         0.0, step=500.0, format="%.2f")
        v_fung = st.number_input("Fungicida",            0.0, step=100.0, format="%.2f")
        v_herb = st.number_input("Herbicida",            0.0, step=100.0, format="%.2f")
        v_inse = st.number_input("Inseticida",           0.0, step=100.0, format="%.2f")
        v_inoc = st.number_input("Inoculantes",          0.0, step=50.0,  format="%.2f")
        v_adju = st.number_input("Adjuvante",            0.0, step=50.0,  format="%.2f")
    with cB:
        st.markdown(f'<small style="color:{_LABEL};font-family:{_FONT}">OPERAÇÕES E SERVIÇOS</small>', unsafe_allow_html=True)
        v_maq  = st.number_input("Op. c/ Máquinas",     0.0, step=200.0, format="%.2f")
        v_seg  = st.number_input("Seguro Agrícola",      0.0, step=100.0, format="%.2f")
        v_tra  = st.number_input("Transporte Externo",   0.0, step=100.0, format="%.2f")
        v_arm  = st.number_input("Armazenagem",          0.0, step=100.0, format="%.2f")
        v_ast  = st.number_input("Assistência Técnica",  0.0, step=50.0,  format="%.2f")
        v_man  = st.number_input("Manutenção Máquinas",  0.0, step=100.0, format="%.2f")
        v_mob  = st.number_input("Mão de Obra",          0.0, step=100.0, format="%.2f")
        v_adm  = st.number_input("Desp. Administrativas",0.0, step=100.0, format="%.2f")
    with cC:
        st.markdown(f'<small style="color:{_LABEL};font-family:{_FONT}">CUSTOS FINANCEIROS</small>', unsafe_allow_html=True)
        v_jur  = st.number_input("Juros / Financiamentos", 0.0, step=100.0, format="%.2f")
        v_dep  = st.number_input(
            "Depreciação (R$/ha)", float(_deprec_ref), step=50.0, format="%.2f",
            help=f"Referência CONAB/Aprosoja: R$ {_deprec_ref:,.2f}/ha. Valor por hectare — será multiplicado pela área.",
        )
        v_out  = st.number_input("Outros Custos",          0.0, step=100.0, format="%.2f")

    c_sub, c_clr = st.columns([5, 1])
    with c_sub:
        submitted = st.form_submit_button("🔍  GERAR RESULTADO", type="primary", use_container_width=True)
    with c_clr:
        cleared = st.form_submit_button("🗑️  LIMPAR", use_container_width=True)

if cleared:
    st.session_state["sim_clear_v"] += 1
    st.rerun()

if submitted:
    erros = []
    if not sim_mun:    erros.append("Município é obrigatório.")
    if sim_area  <= 0: erros.append("Área deve ser maior que zero.")
    if sim_prod  <= 0: erros.append("Produtividade deve ser maior que zero.")

    if erros:
        for e in erros:
            st.error(e)
    else:
        # v_dep é R$/ha (referência CONAB); os demais campos são R$ totais da lavoura
        _dep_total  = v_dep * sim_area
        custo_r = sum([v_sem,v_trat,v_corr,v_fert,v_fung,v_herb,v_inse,v_inoc,v_adju,
                       v_maq,v_seg,v_tra,v_arm,v_ast,v_man,v_mob,v_adm,v_jur,v_out]) + _dep_total
        custo_ha_s  = custo_r / sim_area
        custo_sc_s  = custo_ha_s / sim_prod
        receita_s   = sim_prod * sim_preco
        margem_s    = receita_s - custo_ha_s
        lajida_s    = margem_s + (v_jur / sim_area) + v_dep
        ponto_eq_s  = custo_ha_s / sim_preco if sim_preco else 0

        st.markdown(
            f'<div class="result-banner">'
            f'<p style="color:{_GOLD};font-weight:700;font-size:1rem;'
            f'font-family:{_FONT};margin-bottom:12px">'
            f'Resultado — {sim_mun} · {sim_area:,.0f} ha · {cultura_sel}</p>',
            unsafe_allow_html=True,
        )
        r1, r2, r3, r4, r5, r6 = st.columns(6)
        mc  = _GREEN if margem_s  >= 0 else "#e07878"
        lc  = _GREEN if lajida_s  >= 0 else "#e07878"
        with r1: st.markdown(_kpi("Custo Total / ha",    fmt_brl(custo_ha_s),    "",                  _GOLD),  unsafe_allow_html=True)
        with r2: st.markdown(_kpi("Custo / Saca",        fmt_brl(custo_sc_s),    "R$/sc",             _LIGHT), unsafe_allow_html=True)
        with r3: st.markdown(_kpi("Receita / ha",        fmt_brl(receita_s),     f"{sim_prod:.1f} sc/ha", _LIGHT), unsafe_allow_html=True)
        with r4: st.markdown(_kpi("Margem Líquida / ha", fmt_brl(margem_s),      "",                  mc),     unsafe_allow_html=True)
        with r5: st.markdown(_kpi("LAJIDA / ha",         fmt_brl(lajida_s),      "",                  lc),     unsafe_allow_html=True)
        with r6: st.markdown(_kpi("Ponto de Equilíbrio", f"{ponto_eq_s:.1f} sc/ha", f"Prod.: {sim_prod:.1f} sc/ha", _GOLD), unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

        vals_sim = dict(
            area=sim_area, produtividade=sim_prod, preco=sim_preco,
            sementes=v_sem, trat_semente=v_trat, corretivo=v_corr,
            fertilizante=v_fert, fungicida=v_fung, herbicida=v_herb,
            inseticida=v_inse, inoculantes=v_inoc, adjuvante=v_adju,
            maquinas=v_maq, seguro=v_seg, transporte=v_tra,
            armazenagem=v_arm, assist_tec=v_ast, manutencao=v_man,
            mao_obra=v_mob, desp_admin=v_adm, juros=v_jur, deprec=_dep_total, outros=v_out,
        )
        kpis_sim = dict(
            custo_total_r=custo_r, custo_ha=custo_ha_s, custo_sc=custo_sc_s,
            receita_ha=receita_s, margem_ha=margem_s, ponto_eq=ponto_eq_s,
        )
        try:
            salvar_produtor(sim_mun, vals_sim, kpis_sim)
            ok, err = _push_dados_github()
            if ok:
                st.success(f"✅ Dados salvos e sincronizados com o repositório (aba: {sim_mun})")
            else:
                st.success(f"✅ Dados salvos em **Dados_Produtores.xlsx** (aba: {sim_mun})")
                if err:
                    st.caption(f"⚠ Sincronização GitHub pendente: {err}")
        except Exception as ex:
            st.warning(f"Não foi possível salvar: {ex}")

# ─── FOOTER ───────────────────────────────────────────────────────────────────
st.markdown(
    f'<div class="apro-footer"><img src="data:image/png;base64,{apro_b64}" /></div>',
    unsafe_allow_html=True,
)
st.markdown('<p class="apro-caption">Fonte: Aprosoja/MS · Custo de Produção Safra 2025/2026</p>', unsafe_allow_html=True)
