import streamlit as st
import streamlit.components.v1 as components
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date
import base64
import os
import json as _json

# ─── PATHS ───────────────────────────────────────────────────────────────────
BASE_DIR         = os.path.dirname(os.path.abspath(__file__))
ASSETS           = os.path.join(BASE_DIR, "assets")
EXCEL_DADOS      = os.path.join(BASE_DIR, "Dados_Produtores.xlsx")
SYNC_STATUS_FILE = os.path.join(BASE_DIR, "sync_status.json")

# ─── PAGE CONFIG ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Coleta de Custos – Aprosoja/MS",
    page_icon="🌱",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── IMAGENS ─────────────────────────────────────────────────────────────────
def _b64(path: str) -> str:
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode()

_logo_path  = os.path.join(ASSETS, "Aprosoja-4.png")
_fundo_path = os.path.join(ASSETS, "Fundo da pagina_clean.png")
apro_b64  = _b64(_logo_path)  if os.path.exists(_logo_path)  else ""
fundo_b64 = _b64(_fundo_path) if os.path.exists(_fundo_path) else None

# ─── FONTES ──────────────────────────────────────────────────────────────────
_font_face_css = ""
_font_path = os.path.join(ASSETS, "AprosojaMS.ttf")
if os.path.exists(_font_path):
    _fb64 = _b64(_font_path)
    _font_face_css = (
        "@font-face { font-family: 'AprosojaMS'; "
        f"src: url('data:font/truetype;base64,{_fb64}') format('truetype'); }}"
    )
_FONT = "'AprosojaMS', 'Segoe UI', Tahoma, Geneva, sans-serif"

# ─── CSS ─────────────────────────────────────────────────────────────────────
_fundo_css = ""
if fundo_b64:
    _fundo_css = f"""
  [data-testid="stAppViewContainer"] {{
      background-image:
          linear-gradient(rgba(14,26,18,0.97), rgba(14,26,18,0.97)),
          url("data:image/png;base64,{fundo_b64}") !important;
      background-size: cover !important; background-attachment: fixed !important;
  }}
  [data-testid="stMain"] {{ background: transparent !important; }}
"""

st.markdown(f"""
<style>
  {_font_face_css}
  {_fundo_css}

  html, body, [data-testid="stAppViewContainer"], [data-testid="stMain"],
  [data-testid="stSidebar"], [data-testid="stMarkdownContainer"],
  p, div.stApp, label, h1, h2, h3, h4, h5, h6 {{
      font-family: {_FONT} !important;
  }}

  [data-testid="stAppViewContainer"] {{ background-color: #0e1a12; }}
  [data-testid="stMain"]             {{ background-color: #0e1a12; }}
  .block-container                   {{ padding-top: 2rem !important; }}

  [data-testid="stSidebar"] > div:first-child {{
      background: linear-gradient(180deg,#081410 0%,#0c1e16 60%,#081410 100%) !important;
      border-right: 1px solid #1c3d28 !important;
  }}
  [data-testid="stSidebar"] label, [data-testid="stSidebar"] p,
  [data-testid="stSidebar"] span,   [data-testid="stSidebar"] .stMarkdown {{
      color: #e8f5d0 !important;
  }}
  [data-testid="stSidebar"] hr {{ border-color: #1c3d28 !important; }}
  [data-testid="stSidebar"] [data-testid="stHorizontalBlock"] [data-testid="baseButton-secondary"] {{
      background: #1a3025 !important; color: #c8e6a0 !important;
      border: 1px solid #2d5a3d !important; border-radius: 14px !important;
      font-weight: 600 !important; width: 100% !important;
  }}
  [data-testid="stSidebar"] [data-testid="stHorizontalBlock"] [data-testid="baseButton-primary"] {{
      background: #1a3025 !important; color: #ffffff !important;
      border: 2px solid #4caf50 !important;
      box-shadow: 0 0 10px rgba(76,175,80,0.35) !important;
      border-radius: 14px !important; font-weight: 700 !important; width: 100% !important;
  }}

  [data-testid="stMarkdownContainer"] p,
  [data-testid="stMarkdownContainer"] span,
  .stMarkdown p {{ color: #e0ebe4 !important; }}

  h2, h3 {{
      color: #e8f0eb !important;
      border-left: 5px solid #c8a415;
      padding-left: 10px;
  }}
  hr {{ border-color: #1c3d28 !important; border-width: 1.5px !important; }}

  .apro-header {{
      width:100%; background-color:#2D5416; border-radius:10px; overflow:hidden;
      margin-bottom:1rem; box-shadow:0 3px 10px rgba(0,0,0,.3);
      display:flex; align-items:center; justify-content:center; padding:16px 0;
  }}
  .apro-header img {{ width:50%; max-width:480px; display:block; }}
  .apro-footer {{
      width:100%; background-color:#2D5416; border-radius:8px; overflow:hidden;
      margin-top:0.5rem; box-shadow:0 -2px 6px rgba(0,0,0,.15);
      display:flex; align-items:center; justify-content:center; padding:12px 0;
  }}
  .apro-footer img {{ width:38%; max-width:360px; display:block; }}
  .apro-caption {{ text-align:center; color:#5a8c65; font-size:0.76rem; margin-top:5px; }}

  .sim-header {{
      background: linear-gradient(145deg, #122518 0%, #1a3328 100%);
      border: 1px solid #243f2f; border-radius: 12px;
      padding: 14px 20px; margin-bottom: 14px;
      box-shadow: 0 4px 12px rgba(0,0,0,0.35);
  }}
  .sim-header h3 {{
      color: #e8f0eb !important; border-left: 4px solid #c8a415 !important;
      padding-left: 10px !important; margin: 0 !important; font-size: 1.05rem !important;
  }}
  .result-banner {{
      background: linear-gradient(145deg, #0d2918 0%, #122518 100%);
      border: 1px solid #2d5a3d; border-radius: 12px;
      padding: 18px 20px; margin-top: 14px;
      box-shadow: 0 4px 16px rgba(0,0,0,0.4);
  }}

  [data-testid="stNumberInput"] input,
  [data-testid="stSelectbox"] div {{
      background-color: #1a3025 !important;
      color: #e8f0eb !important;
      border-color: #2d5a3d !important;
  }}
  .stForm [data-testid="stFormSubmitButton"] button {{
      background: #c8a415 !important; color: #0e1a12 !important;
      font-weight: 700 !important; border: none !important;
      border-radius: 8px !important; font-size: 1rem !important;
  }}
  .stForm [data-testid="stFormSubmitButton"] button:hover {{ background: #d4b020 !important; }}
</style>
""" + (f'<div class="apro-header"><img src="data:image/png;base64,{apro_b64}" /></div>'
       if apro_b64 else ""), unsafe_allow_html=True)

# ─── PALETA ───────────────────────────────────────────────────────────────────
_GOLD  = "#c8a415"
_GREEN = "#4caf50"
_LIGHT = "#e8f0eb"
_MUTED = "#7aa88a"
_LABEL = "#6a9978"

# ─── REFERÊNCIAS ─────────────────────────────────────────────────────────────
SOJA_DEPREC           = 325.00
MILHO_DEPREC          = 250.00
_PRECO_SOJA_FALLBACK  = 120.0
_PRECO_MILHO_FALLBACK =  54.0

# ─── KPI CARD ─────────────────────────────────────────────────────────────────
def _kpi(titulo, valor, sub="", cor="#c8a415", sub_cor="#4a6e55"):
    return (
        f'<div style="background:linear-gradient(145deg,#122518 0%,#1a3328 100%);'
        f'border:1px solid #243f2f;border-radius:16px;padding:18px 12px 14px;'
        f'text-align:center;min-height:108px;display:flex;flex-direction:column;'
        f'justify-content:center;box-shadow:0 6px 24px rgba(0,0,0,.45);">'
        f'<div style="color:#6a9978;font-size:.63rem;font-weight:700;'
        f'text-transform:uppercase;letter-spacing:.9px;margin-bottom:7px;'
        f'font-family:{_FONT}">{titulo}</div>'
        f'<div style="font-size:1.45rem;font-weight:800;line-height:1.1;'
        f'color:{cor};font-family:{_FONT}">{valor}</div>'
        f'<div style="color:{sub_cor};font-size:.62rem;margin-top:4px;'
        f'font-family:{_FONT}">{sub}</div>'
        f'</div>'
    )

def fmt_brl(v, dec=2):
    s = f"{v:,.{dec}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"

# ─── MUNICÍPIOS PADRÃO MS ─────────────────────────────────────────────────────
MUNICIPIOS_MS = [
    "Água Clara","Alcinópolis","Amambai","Anastácio","Anaurilândia","Angélica",
    "Antônio João","Aparecida do Taboado","Aquidauana","Aral Moreira","Bandeirantes",
    "Bataguassu","Batayporã","Bela Vista","Bodoquena","Bonito","Brasilândia",
    "Caarapó","Camapuã","Campo Grande","Caracol","Cassilândia","Chapadão do Sul",
    "Corguinho","Coronel Sapucaia","Corumbá","Costa Rica","Coxim","Deodápolis",
    "Dois Irmãos do Buriti","Douradina","Dourados","Eldorado","Fátima do Sul",
    "Figueirão","Glória de Dourados","Guia Lopes da Laguna","Iguatemi","Inocência",
    "Itaporã","Itaquiraí","Ivinhema","Japorã","Jaraguari","Jardim","Jateí","Juti",
    "Ladário","Laguna Carapã","Maracaju","Miranda","Mundo Novo","Naviraí","Nioaque",
    "Nova Alvorada do Sul","Nova Andradina","Novo Horizonte do Sul","Paraíso das Águas",
    "Paranaíba","Paranhos","Pedro Gomes","Ponta Porã","Porto Murtinho",
    "Ribas do Rio Pardo","Rio Brilhante","Rio Negro","Rio Verde de Mato Grosso",
    "Rochedo","Santa Rita do Pardo","São Gabriel do Oeste","Selvíria","Sete Quedas",
    "Sidrolândia","Sonora","Tacuru","Taquarussu","Terenos","Três Lagoas","Vicentina",
]

# ─── COLUNAS EXCEL ────────────────────────────────────────────────────────────
COLUNAS_EXCEL = [
    "Data","Técnico","Município","Área (ha)","Produtividade (sc/ha)","Preço Venda (R$/sc)",
    "Sementes (R$)","Tratamento de Semente (R$)","Corretivo de Solo (R$)",
    "Fertilizante (R$)","Fungicida (R$)","Herbicida (R$)","Inseticida (R$)",
    "Inoculantes (R$)","Adjuvante (R$)","Op. c/ Máquinas (R$)","Seguro Agrícola (R$)",
    "Transporte Externo (R$)","Armazenagem (R$)","Assistência Técnica (R$)",
    "Manutenção Máquinas (R$)","Mão de Obra (R$)","Desp. Administrativas (R$)",
    "Juros/Financiamentos (R$)","Depreciação (R$)","Outros Custos (R$)",
    "CUSTO TOTAL (R$)","CUSTO/ha (R$)","CUSTO/sc (R$)",
    "RECEITA/ha (R$)","MARGEM/ha (R$)","PONTO DE EQUILÍBRIO (sc/ha)",
]

# ─── CARREGAMENTO DE DADOS ────────────────────────────────────────────────────
@st.cache_data
def load_municipios():
    path = os.path.join(BASE_DIR, "Regiões.xlsx")
    if not os.path.exists(path):
        return MUNICIPIOS_MS  # fallback com lista completa de MS
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["Planilha1"]
        munis = sorted({str(row[1]).strip() for row in ws.iter_rows(min_row=2, values_only=True)
                        if row[1] and str(row[1]).strip()})
        wb.close()
        return munis if munis else MUNICIPIOS_MS
    except Exception:
        return MUNICIPIOS_MS

@st.cache_data(ttl=300)
def load_precos_ref():
    path = os.path.join(BASE_DIR, "Preços 2.0.xlsx")
    if not os.path.exists(path):
        return _PRECO_SOJA_FALLBACK, _PRECO_MILHO_FALLBACK
    SOJA_COLS  = [2, 6, 8, 10, 12, 14, 16, 18]
    MILHO_COLS = [4, 7, 9, 11, 13, 15, 17, 19]
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        records = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            sv = [row[c] for c in SOJA_COLS  if c < len(row) and isinstance(row[c], (int, float))]
            mv = [row[c] for c in MILHO_COLS if c < len(row) and isinstance(row[c], (int, float))]
            if sv or mv:
                records.append((sum(sv)/len(sv) if sv else None,
                                sum(mv)/len(mv) if mv else None))
        wb.close()
        records = [(s, m) for s, m in records if s and m]
        if records:
            n  = min(5, len(records))
            ps = round(sum(r[0] for r in records[-n:]) / n, 2)
            pm = round(sum(r[1] for r in records[-n:]) / n, 2)
            return ps, pm
    except Exception:
        pass
    return _PRECO_SOJA_FALLBACK, _PRECO_MILHO_FALLBACK

# ─── SYNC / PERSISTÊNCIA ─────────────────────────────────────────────────────
@st.cache_data(ttl=15)
def _check_online() -> bool:
    import urllib.request
    try:
        req = urllib.request.Request(
            "https://api.github.com",
            headers={"User-Agent": "aprosoja-coleta/1.0"},
        )
        with urllib.request.urlopen(req, timeout=3):
            return True
    except Exception:
        return False

def _load_sync_status() -> dict:
    if os.path.exists(SYNC_STATUS_FILE):
        try:
            with open(SYNC_STATUS_FILE, "r", encoding="utf-8") as f:
                return _json.load(f)
        except Exception:
            pass
    return {"last_sync": None, "pending": False}

def _mark_pending():
    s = _load_sync_status()
    s["pending"] = True
    with open(SYNC_STATUS_FILE, "w", encoding="utf-8") as f:
        _json.dump(s, f, ensure_ascii=False)

def _mark_synced():
    from datetime import datetime as _dt
    with open(SYNC_STATUS_FILE, "w", encoding="utf-8") as f:
        _json.dump({"last_sync": _dt.now().isoformat(), "pending": False}, f, ensure_ascii=False)

def _github_token() -> str:
    token = os.environ.get("GITHUB_TOKEN", "")
    if not token:
        try:
            token = st.secrets.get("GITHUB_TOKEN", "")
        except Exception:
            pass
    return token or ""

def _pull_dados_github() -> bool:
    import urllib.request, urllib.error, json as _j
    repo    = os.environ.get("GITHUB_REPO", "economia2-cyber/coletadecustos")
    token   = _github_token()
    api     = f"https://api.github.com/repos/{repo}/contents/Dados_Produtores.xlsx"
    headers = {"Accept": "application/vnd.github.v3+json"}
    if token:
        headers["Authorization"] = f"token {token}"
    try:
        req = urllib.request.Request(api, headers=headers)
        with urllib.request.urlopen(req, timeout=5) as r:
            data = _j.loads(r.read())
        content = base64.b64decode(data["content"].replace("\n", ""))
        with open(EXCEL_DADOS, "wb") as f:
            f.write(content)
        return True
    except urllib.error.HTTPError as e:
        return e.code == 404
    except Exception:
        return False

def _push_dados_github() -> tuple:
    import urllib.request, urllib.error, json as _j
    token = _github_token()
    repo  = os.environ.get("GITHUB_REPO", "economia2-cyber/coletadecustos")
    if not token:
        return False, "GITHUB_TOKEN não configurado"
    if not os.path.exists(EXCEL_DADOS):
        return False, "Arquivo local não encontrado"
    api     = f"https://api.github.com/repos/{repo}/contents/Dados_Produtores.xlsx"
    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json",
        "Content-Type": "application/json",
    }
    sha = None
    try:
        req = urllib.request.Request(api, headers=headers)
        with urllib.request.urlopen(req, timeout=5) as r:
            sha = _j.loads(r.read()).get("sha")
    except urllib.error.HTTPError as e:
        if e.code != 404:
            return False, f"GitHub API {e.code}"
    except Exception as e:
        return False, str(e)
    with open(EXCEL_DADOS, "rb") as f:
        content_b64 = base64.b64encode(f.read()).decode()
    payload: dict = {"message": f"Coleta — {date.today().isoformat()}", "content": content_b64}
    if sha:
        payload["sha"] = sha
    try:
        body = _j.dumps(payload).encode()
        req  = urllib.request.Request(api, data=body, headers=headers, method="PUT")
        with urllib.request.urlopen(req, timeout=10):
            pass
        return True, ""
    except Exception as e:
        return False, str(e)

def salvar_produtor(municipio, cultura, valores, kpis):
    # Pull só se online — não bloqueia o save local se estiver offline
    if _check_online():
        _pull_dados_github()
    if os.path.exists(EXCEL_DADOS):
        wb = openpyxl.load_workbook(EXCEL_DADOS)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
    nome_aba = f"{municipio} ({cultura})"[:31]
    if nome_aba not in wb.sheetnames:
        ws = wb.create_sheet(nome_aba)
        ws.append(COLUNAS_EXCEL)
        hf  = PatternFill("solid", fgColor="1B4332")
        hft = Font(bold=True, color="FFFFFF")
        for ci in range(1, len(COLUNAS_EXCEL) + 1):
            c = ws.cell(row=1, column=ci)
            c.fill, c.font = hf, hft
            c.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
    else:
        ws = wb[nome_aba]
    ws.append([
        date.today().isoformat(), valores.get("tecnico", ""), municipio,
        valores["area"], valores["produtividade"], valores["preco"],
        valores["sementes"], valores["trat_semente"], valores["corretivo"],
        valores["fertilizante"], valores["fungicida"], valores["herbicida"],
        valores["inseticida"], valores["inoculantes"], valores["adjuvante"],
        valores["maquinas"], valores["seguro"], valores["transporte"],
        valores["armazenagem"], valores["assist_tec"], valores["manutencao"],
        valores["mao_obra"], valores["desp_admin"],
        valores["juros"], valores["deprec"], valores["outros"],
        kpis["custo_total_r"], kpis["custo_ha"], kpis["custo_sc"],
        kpis["receita_ha"], kpis["margem_ha"], kpis["ponto_eq"],
    ])
    wb.save(EXCEL_DADOS)

def _auto_sync_pending() -> tuple:
    if not _load_sync_status().get("pending"):
        return False, ""
    if not _check_online():
        return False, "offline"
    ok, err = _push_dados_github()
    if ok:
        _mark_synced()
        return True, ""
    return False, err or "falha ao sincronizar"


def importar_dados_campo() -> tuple:
    """Lê dados_campo.json do GitHub, grava novas entradas no Excel e limpa o arquivo remoto."""
    import urllib.request, urllib.error, json as _j
    token = _github_token()
    repo  = os.environ.get("GITHUB_REPO", "economia2-cyber/coletadecustos")
    if not token:
        return 0, "GITHUB_TOKEN não configurado"

    api     = f"https://api.github.com/repos/{repo}/contents/dados_campo.json"
    headers = {"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json"}

    # Baixa dados_campo.json
    try:
        req = urllib.request.Request(api, headers=headers)
        with urllib.request.urlopen(req, timeout=5) as r:
            meta = _j.loads(r.read())
        sha     = meta["sha"]
        entries = _j.loads(base64.b64decode(meta["content"].replace("\n", "")))
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return 0, "Nenhum dado de campo encontrado no repositório"
        return 0, f"GitHub API {e.code}"
    except Exception as e:
        return 0, str(e)

    if not entries:
        return 0, "Arquivo de campo está vazio"

    # Grava no Excel
    if os.path.exists(EXCEL_DADOS):
        wb = openpyxl.load_workbook(EXCEL_DADOS)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    salvos = 0
    for e in entries:
        nome_aba = f"{e.get('municipio','?')} ({e.get('cultura','?')})"[:31]
        if nome_aba not in wb.sheetnames:
            ws = wb.create_sheet(nome_aba)
            ws.append(COLUNAS_EXCEL)
            hf  = PatternFill("solid", fgColor="1B4332")
            hft = Font(bold=True, color="FFFFFF")
            for ci in range(1, len(COLUNAS_EXCEL) + 1):
                c = ws.cell(row=1, column=ci)
                c.fill, c.font = hf, hft
                c.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"
        else:
            ws = wb[nome_aba]
        ws.append([
            e.get("data",""), e.get("tecnico",""), e.get("municipio",""),
            e.get("area",0), e.get("produtividade",0), e.get("preco",0),
            e.get("sementes",0), e.get("trat_semente",0), e.get("corretivo",0),
            e.get("fertilizante",0), e.get("fungicida",0), e.get("herbicida",0),
            e.get("inseticida",0), e.get("inoculantes",0), e.get("adjuvante",0),
            e.get("maquinas",0), e.get("seguro",0), e.get("transporte",0),
            e.get("armazenagem",0), e.get("assist_tec",0), e.get("manutencao",0),
            e.get("mao_obra",0), e.get("desp_admin",0),
            e.get("juros",0), e.get("deprec",0), e.get("outros",0),
            e.get("custo_total_r",0), e.get("custo_ha",0), e.get("custo_sc",0),
            e.get("receita_ha",0), e.get("margem_ha",0), e.get("ponto_eq",0),
        ])
        salvos += 1
    wb.save(EXCEL_DADOS)

    # Limpa dados_campo.json no GitHub (substitui por array vazio)
    try:
        import json as _jj
        body = _jj.dumps({
            "message": f"Importado — {date.today().isoformat()}",
            "content": base64.b64encode(b"[]").decode(),
            "sha": sha,
        }).encode()
        req = urllib.request.Request(api, data=body, headers={**headers, "Content-Type": "application/json"}, method="PUT")
        with urllib.request.urlopen(req, timeout=10):
            pass
    except Exception:
        pass  # não crítico

    # Sincroniza Excel atualizado
    _push_dados_github()
    return salvos, ""

# ─── SIDEBAR ─────────────────────────────────────────────────────────────────
with st.sidebar:
    if apro_b64:
        st.markdown(
            f'<img src="data:image/png;base64,{apro_b64}" '
            f'style="width:100%;border-radius:8px;margin-bottom:12px;">',
            unsafe_allow_html=True,
        )
    st.markdown("---")

    _sel_now = st.session_state.get("_cultura_sel", "Soja IPRO")
    _c1, _c2 = st.columns(2)
    with _c1:
        if st.button("🌱 SOJA", use_container_width=True, key="btn_soja",
                     type="primary" if _sel_now == "Soja IPRO" else "secondary"):
            st.session_state["_cultura_sel"] = "Soja IPRO"
            st.rerun()
    with _c2:
        if st.button("🌽 MILHO", use_container_width=True, key="btn_milho",
                     type="primary" if _sel_now == "Milho" else "secondary"):
            st.session_state["_cultura_sel"] = "Milho"
            st.rerun()

    _lbl_sel = "SOJA" if _sel_now == "Soja IPRO" else "MILHO"
    components.html(f"""
<script>
(function(){{
  var SEL = "{_lbl_sel}";
  function applyStyles(){{
    var sb = window.parent.document.querySelector('[data-testid="stSidebar"]');
    if(!sb) return false;
    var hb = sb.querySelector('[data-testid="stHorizontalBlock"]');
    if(!hb) return false;
    var btns = hb.querySelectorAll('button');
    if(!btns.length) return false;
    btns.forEach(function(b){{
      var isSel = b.textContent.indexOf(SEL) !== -1;
      b.style.setProperty('background','#3a6b1a','important');
      b.style.setProperty('border-radius','14px','important');
      b.style.setProperty('font-weight', isSel?'700':'600','important');
      b.style.setProperty('color', isSel?'#ffffff':'#d4edaa','important');
      b.style.setProperty('border', isSel?'3px solid #e74c3c':'2px solid #5a9b2a','important');
      b.style.setProperty('box-shadow', isSel?'0 0 10px rgba(231,76,60,0.4)':'none','important');
    }});
    return true;
  }}
  if(!applyStyles()){{
    var t=0,iv=setInterval(function(){{if(applyStyles()||++t>20)clearInterval(iv);}},100);
  }}
}})();
</script>
""", height=0)

    cultura_sel = _sel_now
    st.markdown("---")
    st.markdown(
        '<div style="background:linear-gradient(135deg,#1a4a28,#2d6b3a);'
        'border:2px solid #c8a415;border-radius:12px;padding:10px 14px;'
        'text-align:center;box-shadow:0 0 10px rgba(200,164,21,0.25);">'
        '<span style="color:#c8a415;font-weight:700;font-size:0.9rem;'
        'letter-spacing:1px;text-transform:uppercase;">📋 Coleta de Custos</span>'
        '</div>',
        unsafe_allow_html=True,
    )

    # ── Importar dados dos tablets ──────────────────────────────────────────
    st.markdown("---")
    st.markdown(
        '<span style="color:#6a9978;font-size:0.70rem;font-weight:700;'
        'text-transform:uppercase;letter-spacing:0.8px;">📱 Dados dos Tablets</span>',
        unsafe_allow_html=True,
    )
    if st.button("⬇️ Importar do Campo", use_container_width=True, key="btn_importar"):
        with st.spinner("Importando dados_campo.json do GitHub..."):
            _n, _err = importar_dados_campo()
        if _n > 0:
            st.success(f"✅ {_n} entrada(s) importada(s) para o Excel!")
        elif _err:
            st.error(f"❌ {_err}")
        else:
            st.info("Nenhum dado novo encontrado.")

    # ── Status de sincronização ─────────────────────────────────────────────
    st.markdown("---")
    _sync_st = _load_sync_status()
    _online  = _check_online()

    if _sync_st.get("pending"):
        if _online:
            st.markdown(
                '<div style="background:#1a2a0e;border:1px solid #c8a415;border-radius:8px;'
                'padding:8px 10px;text-align:center;margin-bottom:6px;">'
                '<span style="color:#c8a415;font-size:0.70rem;font-weight:700;">'
                '⚠ Dados pendentes de sincronização</span></div>',
                unsafe_allow_html=True,
            )
            if st.button("🔄 Sincronizar agora", use_container_width=True, key="btn_sync_now"):
                _ok, _err = _push_dados_github()
                if _ok:
                    _mark_synced()
                    _check_online.clear()
                    st.rerun()
                else:
                    st.error(f"Falha: {_err}")
        else:
            st.markdown(
                '<div style="background:#1a1010;border:1px solid #8b3333;border-radius:8px;'
                'padding:8px 10px;text-align:center;">'
                '<span style="color:#e07878;font-size:0.70rem;font-weight:700;">'
                '📴 Offline</span><br>'
                '<span style="color:#b06060;font-size:0.63rem;">'
                'Dados salvos localmente.<br>Sincronizará ao reconectar.</span></div>',
                unsafe_allow_html=True,
            )
    else:
        _last = _sync_st.get("last_sync")
        if _last:
            from datetime import datetime as _dt_cls
            _dt_str = _dt_cls.fromisoformat(_last).strftime("%d/%m/%Y %H:%M")
            st.markdown(
                f'<div style="background:#0a1e10;border:1px solid #1e4a2a;border-radius:8px;'
                f'padding:8px 10px;text-align:center;">'
                f'<span style="color:#5ed67f;font-size:0.70rem;font-weight:700;">'
                f'✅ Sincronizado</span><br>'
                f'<span style="color:#7aa88a;font-size:0.62rem;">{_dt_str}</span></div>',
                unsafe_allow_html=True,
            )
        else:
            _sc = "#5ed67f" if _online else "#6a9978"
            _si = "🌐" if _online else "📴"
            _st_txt = "Online" if _online else "Offline"
            st.markdown(
                f'<div style="background:#0a1e10;border:1px solid #1e3a20;border-radius:8px;'
                f'padding:6px 10px;text-align:center;">'
                f'<span style="color:{_sc};font-size:0.70rem;">{_si} {_st_txt}</span></div>',
                unsafe_allow_html=True,
            )

    st.caption("Aprosoja/MS · Safra 2025/2026")

# ─── AUTO-SYNC NA INICIALIZAÇÃO ───────────────────────────────────────────────
if "startup_sync_done" not in st.session_state:
    st.session_state["startup_sync_done"] = True
    _synced, _sync_err = _auto_sync_pending()
    if _synced:
        st.success("✅ Dados coletados offline foram sincronizados automaticamente com o repositório.")
    elif _sync_err and _sync_err != "offline":
        st.warning(f"⚠ Sincronização automática falhou: {_sync_err}")

# ─── CONTEÚDO ─────────────────────────────────────────────────────────────────
_cultura_display = cultura_sel.replace(" IPRO", "")
st.markdown(
    f'<h1 style="color:#e8f0eb;font-size:1.85rem;font-weight:800;'
    f'letter-spacing:2px;margin:8px 0 2px 0;line-height:1.2;font-family:{_FONT}">'
    f'COLETA DE CUSTOS <span style="color:{_GOLD}">{_cultura_display.upper()}</span>'
    f' <span style="font-size:1.2rem;font-weight:400;color:{_MUTED}">| Mato Grosso do Sul</span>'
    f'</h1>',
    unsafe_allow_html=True,
)
st.caption("Safra 2026/2027. Aprosoja/MS — Informe os custos reais da sua lavoura")
st.divider()

# ─── FORMULÁRIO DE COLETA ─────────────────────────────────────────────────────
st.markdown('<div class="sim-header"><h3>📋 Informe os Dados da Sua Lavoura</h3></div>', unsafe_allow_html=True)
st.markdown(
    f'<p style="color:{_MUTED};font-size:0.85rem;font-family:{_FONT};margin-bottom:12px">'
    f'Preencha os valores gastos em cada item da lavoura. '
    f'<b style="color:#e8f0eb">Campos obrigatórios:</b> Técnico, Município, Área e Produtividade.</p>',
    unsafe_allow_html=True,
)

_municipios         = load_municipios()
_preco_soja, _preco_milho = load_precos_ref()
_preco_default      = _preco_soja if cultura_sel == "Soja IPRO" else _preco_milho
_deprec_ref         = SOJA_DEPREC if cultura_sel == "Soja IPRO" else MILHO_DEPREC

if "sim_clear_v" not in st.session_state:
    st.session_state["sim_clear_v"] = 0

with st.form(f"coleta_form_{st.session_state['sim_clear_v']}"):
    rc1, rc2 = st.columns([2, 2])
    with rc1:
        sim_tecnico = st.text_input("Técnico *", placeholder="Digite seu nome completo")
    with rc2:
        if _municipios:
            sim_mun = st.selectbox("Município *", [""] + _municipios)
        else:
            sim_mun = st.text_input("Município *")

    r1c1, r1c2, r1c3 = st.columns(3)
    with r1c1:
        sim_area  = st.number_input("Área (ha) *",             min_value=0.0, value=0.0,  step=10.0, format="%.1f")
    with r1c2:
        sim_prod  = st.number_input("Produtividade (sc/ha) *", min_value=0.0, value=0.0,  step=1.0,  format="%.1f")
    with r1c3:
        sim_preco = st.number_input("Preço de venda (R$/sc)",  min_value=0.0, value=float(_preco_default), step=1.0, format="%.2f")

    st.markdown(
        f'<p style="color:{_MUTED};font-size:0.82rem;font-family:{_FONT};margin:8px 0 4px">'
        f'Informe o valor <b style="color:#e8f0eb">total gasto</b> em cada item (R$):</p>',
        unsafe_allow_html=True,
    )

    cA, cB, cC = st.columns(3)
    with cA:
        st.markdown(f'<small style="color:{_LABEL};font-family:{_FONT}">INSUMOS DA LAVOURA</small>', unsafe_allow_html=True)
        v_sem  = st.number_input("Sementes",              0.0, step=100.0, format="%.2f")
        v_trat = st.number_input("Tratamento de Semente", 0.0, step=50.0,  format="%.2f")
        v_corr = st.number_input("Corretivo de Solo",     0.0, step=100.0, format="%.2f")
        v_fert = st.number_input("Fertilizante",          0.0, step=500.0, format="%.2f")
        v_fung = st.number_input("Fungicida",             0.0, step=100.0, format="%.2f")
        v_herb = st.number_input("Herbicida",             0.0, step=100.0, format="%.2f")
        v_inse = st.number_input("Inseticida",            0.0, step=100.0, format="%.2f")
        v_inoc = st.number_input("Inoculantes",           0.0, step=50.0,  format="%.2f")
        v_adju = st.number_input("Adjuvante",             0.0, step=50.0,  format="%.2f")
    with cB:
        st.markdown(f'<small style="color:{_LABEL};font-family:{_FONT}">OPERAÇÕES E SERVIÇOS</small>', unsafe_allow_html=True)
        v_maq  = st.number_input("Op. c/ Máquinas",       0.0, step=200.0, format="%.2f")
        v_seg  = st.number_input("Seguro Agrícola",        0.0, step=100.0, format="%.2f")
        v_tra  = st.number_input("Transporte Externo",     0.0, step=100.0, format="%.2f")
        v_arm  = st.number_input("Armazenagem",            0.0, step=100.0, format="%.2f")
        v_ast  = st.number_input("Assistência Técnica",    0.0, step=50.0,  format="%.2f")
        v_man  = st.number_input("Manutenção Máquinas",    0.0, step=100.0, format="%.2f")
        v_mob  = st.number_input("Mão de Obra",            0.0, step=100.0, format="%.2f")
        v_adm  = st.number_input("Desp. Administrativas",  0.0, step=100.0, format="%.2f")
    with cC:
        st.markdown(f'<small style="color:{_LABEL};font-family:{_FONT}">CUSTOS FINANCEIROS</small>', unsafe_allow_html=True)
        v_jur = st.number_input("Juros / Financiamentos",  0.0, step=100.0, format="%.2f")
        v_dep = st.number_input(
            "Depreciação (R$/ha)", float(_deprec_ref), step=50.0, format="%.2f",
            help=f"Referência CONAB/Aprosoja: R$ {_deprec_ref:,.2f}/ha. Será multiplicado pela área.",
        )
        v_out = st.number_input("Outros Custos",           0.0, step=100.0, format="%.2f")

    c_sub, c_clr = st.columns([5, 1])
    with c_sub:
        submitted = st.form_submit_button("🔍  CALCULAR E SALVAR", type="primary", use_container_width=True)
    with c_clr:
        cleared = st.form_submit_button("🗑️  LIMPAR", use_container_width=True)

if cleared:
    st.session_state["sim_clear_v"] += 1
    st.rerun()

if submitted:
    erros = []
    if not sim_tecnico.strip(): erros.append("Preencha o campo **Técnico**.")
    if not sim_mun:             erros.append("Selecione o **Município**.")
    if sim_area  <= 0:          erros.append("**Área** deve ser maior que zero.")
    if sim_prod  <= 0:          erros.append("**Produtividade** deve ser maior que zero.")

    if erros:
        for e in erros:
            st.error(e)
    else:
        _dep_total = v_dep * sim_area
        custo_r    = sum([v_sem, v_trat, v_corr, v_fert, v_fung, v_herb,
                          v_inse, v_inoc, v_adju, v_maq, v_seg, v_tra,
                          v_arm, v_ast, v_man, v_mob, v_adm, v_jur, v_out]) + _dep_total
        custo_ha_s = custo_r / sim_area
        custo_sc_s = custo_ha_s / sim_prod
        receita_s  = sim_prod * sim_preco
        margem_s   = receita_s - custo_ha_s
        lajida_s   = margem_s + (v_jur / sim_area) + v_dep
        ponto_eq_s = custo_ha_s / sim_preco if sim_preco else 0

        st.markdown(
            f'<div class="result-banner">'
            f'<p style="color:{_GOLD};font-weight:700;font-size:1rem;'
            f'font-family:{_FONT};margin-bottom:12px">'
            f'Resultado — {sim_mun} · {sim_area:,.0f} ha · {cultura_sel}</p>',
            unsafe_allow_html=True,
        )
        r1, r2, r3, r4, r5, r6 = st.columns(6)
        mc = _GREEN if margem_s >= 0 else "#e07878"
        lc = _GREEN if lajida_s >= 0 else "#e07878"
        with r1: st.markdown(_kpi("Custo Total / ha",    fmt_brl(custo_ha_s), "",                       _GOLD),  unsafe_allow_html=True)
        with r2: st.markdown(_kpi("Custo / Saca",        fmt_brl(custo_sc_s), "R$/sc",                  _LIGHT), unsafe_allow_html=True)
        with r3: st.markdown(_kpi("Receita / ha",        fmt_brl(receita_s),  f"{sim_prod:.1f} sc/ha",  _LIGHT), unsafe_allow_html=True)
        with r4: st.markdown(_kpi("Margem Líquida / ha", fmt_brl(margem_s),   "",                       mc),     unsafe_allow_html=True)
        with r5: st.markdown(_kpi("LAJIDA / ha",         fmt_brl(lajida_s),   "",                       lc),     unsafe_allow_html=True)
        with r6: st.markdown(_kpi("Ponto de Equilíbrio", f"{ponto_eq_s:.1f} sc/ha",
                                  f"Prod.: {sim_prod:.1f} sc/ha", _GOLD), unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

        vals = dict(
            tecnico=sim_tecnico.strip(),
            area=sim_area, produtividade=sim_prod, preco=sim_preco,
            sementes=v_sem, trat_semente=v_trat, corretivo=v_corr,
            fertilizante=v_fert, fungicida=v_fung, herbicida=v_herb,
            inseticida=v_inse, inoculantes=v_inoc, adjuvante=v_adju,
            maquinas=v_maq, seguro=v_seg, transporte=v_tra,
            armazenagem=v_arm, assist_tec=v_ast, manutencao=v_man,
            mao_obra=v_mob, desp_admin=v_adm, juros=v_jur, deprec=_dep_total, outros=v_out,
        )
        kpis = dict(
            custo_total_r=custo_r, custo_ha=custo_ha_s, custo_sc=custo_sc_s,
            receita_ha=receita_s, margem_ha=margem_s, ponto_eq=ponto_eq_s,
        )
        try:
            salvar_produtor(sim_mun, _cultura_display, vals, kpis)
            if _check_online():
                ok, err = _push_dados_github()
                if ok:
                    _mark_synced()
                    st.success(f"✅ Dados de **{sim_tecnico.strip()}** salvos e sincronizados! (aba: {sim_mun})")
                else:
                    _mark_pending()
                    st.success(f"✅ Dados de **{sim_tecnico.strip()}** salvos localmente. (aba: {sim_mun})")
                    if err:
                        st.caption(f"⚠ Sincronização pendente: {err}")
            else:
                _mark_pending()
                st.success(f"✅ Dados de **{sim_tecnico.strip()}** salvos localmente. (aba: {sim_mun})")
                st.info("📴 Sem conexão — sincronizará automaticamente ao reconectar.")
        except Exception as ex:
            st.error(f"❌ Erro ao salvar: {ex}")

# ─── FOOTER ───────────────────────────────────────────────────────────────────
st.divider()
if apro_b64:
    st.markdown(
        f'<div class="apro-footer"><img src="data:image/png;base64,{apro_b64}" /></div>',
        unsafe_allow_html=True,
    )
st.markdown('<p class="apro-caption">Fonte: Aprosoja/MS · Custo de Produção Safra 2025/2026</p>', unsafe_allow_html=True)
